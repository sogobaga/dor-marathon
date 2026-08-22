# 城市探索關主：xlsx 權威名單(20260714 名稱差異化優化版) vs 正式庫 explore_bosses 差異報告。
# 只讀不寫。執行：python scripts/card-upload/diff_bosses.py
import re
import ssl
import sys
from collections import Counter

import truststore

truststore.inject_into_ssl()
import pg8000.native
from openpyxl import load_workbook

sys.stdout.reconfigure(encoding="utf-8")
REPO = "c:/Project Dev/04_Online Marathon Project"
XLSX = f"{REPO}/source/data/files/DOR城市探索_關主名稱差異化優化版_20260714.xlsx"

# --- xlsx ---
wb = load_workbook(XLSX, read_only=True)
main_rows = list(wb["打卡地點整合V2V3"].iter_rows(min_row=2, values_only=True))
talk_rows = list(wb["關主挑戰對話"].iter_rows(min_row=2, values_only=True))
main = {}  # code -> dict
for r in main_rows:
    if not r[0]:
        continue
    code = str(r[0]).strip()
    main[code] = {
        "place": str(r[4] or "").strip(),
        "stars": int(r[8] or 0),
        "lat": float(r[6]) if r[6] not in (None, "") else None,
        "lng": float(r[7]) if r[7] not in (None, "") else None,
        "radius": int(r[13]) if r[13] not in (None, "") else None,
    }
# ⚠️ 關主名的權威來源＝主表「打卡地點整合V2V3」的「關主名」欄（使用者維護的主檔）；
# 「關主挑戰對話」表的關主名欄曾出現與主表不一致的殘留（DOR-TRK-NTP-011 竹狐 vs 小月，2026-08-22
# 誤用對話表導致誤改一筆），對話表僅供對話內容參考，名稱/性別/年齡一律以主表為準。
hdr = [str(c) for c in next(wb["打卡地點整合V2V3"].iter_rows(max_row=1, values_only=True))]
qi, gi, ai = hdr.index("關主名"), hdr.index("性別"), hdr.index("年齡")
talk = {}
for r in main_rows:
    if not r[0]:
        continue
    talk[str(r[0]).strip()] = {"name": str(r[qi] or "").strip(), "gender": str(r[gi] or "").strip(), "age": r[ai]}

print(f"xlsx 主表 {len(main)} 筆 / 對話表 {len(talk)} 筆")

# --- DB ---
url = ""
for line in open(f"{REPO}/.env", encoding="utf-8-sig"):
    line = line.strip()
    if line.startswith("DATABASE_URL="):
        url = line.split("=", 1)[1].replace("-pooler", "")
m = re.match(r"postgres(?:ql)?://([^:]+):([^@]+)@([^/]+)/([^?]+)", url)
user, pwd, host, db = m.groups()
ctx = truststore.SSLContext(ssl.PROTOCOL_TLS_CLIENT)
con = pg8000.native.Connection(user, host=host, database=db, password=pwd, ssl_context=ctx)
rows = con.run("SELECT code, name, difficulty_stars, place, lat, lng, radius_m, enabled, gender, age FROM explore_bosses")
con.close()
dbmap = {}
for code, name, stars, place, lat, lng, radius, enabled, gender, age in rows:
    if code:
        dbmap[code.strip()] = {
            "name": (name or "").strip(), "stars": stars or 0, "place": (place or "").strip(),
            "lat": lat, "lng": lng, "radius": radius, "enabled": enabled, "gender": (gender or "").strip(), "age": age,
        }

print(f"DB {len(dbmap)} 筆")

# --- 差異 ---
star_diff, name_diff, place_diff, coord_diff = [], [], [], []
for code, x in main.items():
    d = dbmap.get(code)
    if not d:
        continue
    if d["stars"] != x["stars"]:
        star_diff.append((code, d["stars"], x["stars"]))
    t = talk.get(code)
    if t and t["name"] and d["name"] != t["name"]:
        name_diff.append((code, d["name"], t["name"]))
    if x["place"] and d["place"] and x["place"] not in d["place"] and d["place"] not in x["place"]:
        place_diff.append((code, d["place"][:20], x["place"][:20]))
    if x["lat"] is not None and d["lat"] is not None and (abs(float(d["lat"]) - x["lat"]) > 0.001 or abs(float(d["lng"]) - x["lng"]) > 0.001):
        coord_diff.append((code, f"{d['lat']},{d['lng']}", f"{x['lat']},{x['lng']}"))

xlsx_missing_in_db = [c for c in main if c not in dbmap]
db_not_in_xlsx = [c for c in dbmap if c not in main]
pfx = Counter(re.match(r"^([A-Z]+-[A-Z]+)", c).group(1) if re.match(r"^([A-Z]+-[A-Z]+)", c) else "?" for c in db_not_in_xlsx)

print(f"\n== 星等差異 {len(star_diff)} 筆（樣本10）==")
for c, a, b in star_diff[:10]:
    print(f"  {c}: DB {a} → xlsx {b}")
print(f"\n== 關主名差異 {len(name_diff)} 筆（樣本10）==")
for c, a, b in name_diff[:10]:
    print(f"  {c}: {a} → {b}")
print(f"\n== 地點名差異 {len(place_diff)} 筆（樣本5）==")
for row in place_diff[:5]:
    print("  ", row)
print(f"\n== 座標差異(>0.001度) {len(coord_diff)} 筆（樣本5）==")
for row in coord_diff[:5]:
    print("  ", row)
print(f"\n== xlsx 有但 DB 沒有 {len(xlsx_missing_in_db)} 筆 ==")
for c in xlsx_missing_in_db[:10]:
    print("  ", c)
print(f"\n== DB 有但 xlsx 沒有 {len(db_not_in_xlsx)} 筆，前綴分布 ==")
for p, n in pfx.most_common(12):
    print(f"  {p}: {n}")
